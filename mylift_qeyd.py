import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJfaWQiOiI2OTgwNDQ4YjY0ZWM3MjgzYjNmZDhiNDUiLCJpYXQiOjE3ODAyOTQ1NjF9.OaFZc6bUKQJhvZgn0c1vFhioIOp5FNAKujqc3kh8r0I"

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/json",
    "Accept-Language": "az",
    "Origin": "https://platform.mylift.az",
    "Referer": "https://platform.mylift.az/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

# ── 1. Bütün səhifələri çək ──────────────────────────────────────────────────
all_lifts = []
page = 1

while True:
    resp = requests.get(
        "https://api.mylift.az/lifts",
        headers=HEADERS,
        params={"limit": 50, "page": page,
                "status": "rn_onreview_government",
                "sort": "createdAt", "order": "desc"}
    )
    resp.raise_for_status()
    raw = resp.json()

    items = raw["data"]["items"]
    page_info = raw["data"].get("pageInfo", {})
    total_pages = page_info.get("totalPages", 1)

    all_lifts.extend(items)
    print(f"Səhifə {page}/{total_pages}: {len(items)} lift (cəmi: {len(all_lifts)})")

    if page >= total_pages:
        break
    page += 1

print(f"\nCəmi {len(all_lifts)} lift əldə edildi.")

# ── 2. DataFrame ─────────────────────────────────────────────────────────────
rows = []
for lift in all_lifts:
    building = lift.get("building", {}) or {}
    rows.append({
        "Lift №":               lift.get("publicId", ""),
        "Model":                lift.get("model", ""),
        "İstehsalçı":           lift.get("manufacturer", ""),
        "Zavod nömrəsi":        lift.get("factoryNumber", ""),
        "Tip":                  lift.get("type", ""),
        "Bina adı":             building.get("title", ""),
        "Ünvan":                building.get("address", ""),
        "İstehsal tarixi":      lift.get("manufactureDate", "")[:10] if lift.get("manufactureDate") else "",
        "Servis müddəti (il)":  lift.get("servicePeriod", ""),
        "Yaradılma tarixi":     lift.get("createdAt", "")[:10] if lift.get("createdAt") else "",
        "Passport":             lift.get("passportUrl", ""),
    })

df = pd.DataFrame(rows)

# ── 3. Excel ─────────────────────────────────────────────────────────────────
output_file = "mylift_onreview_government.xlsx"
df.to_excel(output_file, index=False, sheet_name="Liftlər")

wb = load_workbook(output_file)
ws = wb["Liftlər"]

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

for cell in ws[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[1].height = 30

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
    for cell in row:
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        cell.font = Font(name="Arial", size=10)

col_widths = [8, 14, 28, 18, 12, 28, 28, 38, 14, 18, 18, 18, 50]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# Xülasə sheet
ws2 = wb.create_sheet("Xülasə")
ws2["A1"], ws2["B1"] = "Məlumat", "Dəyər"
ws2["A2"], ws2["B2"] = "Ümumi lift sayı", len(df)
ws2["A3"], ws2["B3"] = "Status", "rn_onreview_government"
ws2["A4"], ws2["B4"] = "Çəkilmə tarixi", pd.Timestamp.now().strftime("%d.%m.%Y %H:%M")
for cell in ws2["A1:B1"][0]:
    cell.font = Font(bold=True, color="FFFFFF", name="Arial")
    cell.fill = PatternFill("solid", start_color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 30

wb.save(output_file)
print(f"✅ Excel hazırlandı: {output_file}")